import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import shutil
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "B端精准搜索关键词库"

# Ensure grid lines are visible
ws.views.sheetView[0].showGridLines = True

# Title Block
ws.merge_cells('A1:F1')
title_cell = ws['A1']
title_cell.value = '📊 今遇莨缘 · B端精准引流搜索关键词库'
title_cell.font = Font(name='微软雅黑', size=16, bold=True, color='1E3A8A')
title_cell.alignment = Alignment(horizontal='left', vertical='center')

ws.merge_cells('A2:F2')
sub_cell = ws['A2']
sub_cell.value = '表格一：B端精准搜索关键词库（主要投放 / 抓取词）'
sub_cell.font = Font(name='微软雅黑', size=12, bold=True, color='374151')
sub_cell.alignment = Alignment(horizontal='left', vertical='center')

ws.row_dimensions[1].height = 35
ws.row_dimensions[2].height = 25
ws.row_dimensions[3].height = 10

# Table Headers
headers = ['序号', '关键词分类', '目标搜索关键词', '目标客群画像', '搜索意图与商业价值', '推荐优先级']

ws.row_dimensions[4].height = 28
header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid') # Dark Slate
header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)

header_border = Border(
    left=Side(style='thin', color='475569'),
    right=Side(style='thin', color='475569'),
    top=Side(style='thin', color='475569'),
    bottom=Side(style='thin', color='475569')
)

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_idx, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = header_border

# Data Rows exactly matching prompt image & extended B-end keyword strategy
data = [
    ('01', '源头采购与大货批发', '香云纱源头大厂', '寻找源头工厂直供的买手店主理人', '寻找一级供应链，具备极高采购意向', 'P0（最高）'),
    ('02', '源头采购与大货批发', '香云纱大货批发', '准备大批量进货的二批商 / 国风品牌', '寻求大货配额，客单价高，天然剔除散客', 'P0（最高）'),
    ('03', '源头采购与大货批发', '香云纱买手店进货渠道', '独立买手店 / 精品女装店主理人', '寻求 B 端独家供货渠道，避免大众同质化', 'P0（最高）'),
    ('04', '源头采购与大货批发', '高端国风面料供应链', '原创服装设计师 / 品牌采购总监', '寻找稳定、具备绝对审美门槛的长期供应商', 'P0（最高）'),
    ('05', '源头采购与大货批发', '香云纱高定面料供货', 'Haute Couture 高定工作室 / 私域主理人', '追求极致面料品质，具备极高终端溢价空间', 'P1（高）'),
    ('06', '源头采购与大货批发', '香云纱样卡索样 / 面料调样', '有新季选款/做样衣需求的设计师', '直接产生“拿样-下单”转化，意向极其强烈', 'P0（最高）'),
    ('07', '地理与展厅看样', '深圳南油香云纱展厅', '华南或全国前往南油选货的买手', '拟线下看样实操，到店转化率极高', 'P0（最高）'),
    ('08', '地理与展厅看样', '南油名品馆 103', '明确知晓南油货源地的高端采购', '直达物理展厅，精准度 100% 的行业熟客', 'P0（最高）'),
    ('09', '面料工艺与品类', '龟裂纹香云纱现货批发', '寻找特色稀缺面料的独立设计师', '针对特定工艺品类，寻货目的明确', 'P1（高）'),
    ('10', '面料工艺与品类', '全手晒薯莨香云纱供货', '强调非遗传统工艺的高端国风品牌', '关注极致天然手工工艺与故事赋能', 'P1（高）'),
    ('11', '面料工艺与品类', '桑蚕丝香云纱匹布直供', '中高端女装工厂 / 订制工坊', '原材料匹布采购，用料量大且复购率高', 'P0（最高）'),
    ('12', '商业合作与定制', '香云纱小批量柔性定制', '新锐国风品牌 / 私域爆款主理人', '低门槛小出货，建立长期合作契机', 'P1（高）'),
    ('13', '商业合作与定制', '香云纱贴牌代工 OEM', '无自建工厂但有渠道资源的品牌商', '寻求一站式代工与高品质成品交付', 'P0（最高）'),
    ('14', '场景痛点与组货', '新中式买手店香云纱组货', '连锁买手店采购 / 集合店主理人', '解决组货选品难题，提升店面货品档次', 'P1（高）'),
    ('15', '品牌防守与对标', '今遇莨缘香云纱源头', '已经通过口碑或推荐获知品牌的精准客群', '品牌词防守与直接签约转化', 'P0（最高）'),
]

row_alt_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
row_white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

# Priority styling
p0_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid') # Soft Red
p0_font = Font(name='微软雅黑', size=10, bold=True, color='991B1B')

p1_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid') # Soft Amber
p1_font = Font(name='微软雅黑', size=10, bold=True, color='92400E')

p2_fill = PatternFill(start_color='E0F2FE', end_color='E0F2FE', fill_type='solid') # Soft Blue
p2_font = Font(name='微软雅黑', size=10, bold=True, color='075985')

font_main = Font(name='微软雅黑', size=10, color='1F2937')
font_kw = Font(name='微软雅黑', size=10, bold=True, color='0F172A')

for row_idx, row_data in enumerate(data, 5):
    ws.row_dimensions[row_idx].height = 26
    bg_fill = row_alt_fill if row_idx % 2 == 1 else row_white_fill
    
    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = bg_fill
        cell.border = thin_border
        
        if col_idx == 1:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = font_main
        elif col_idx == 2:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = font_main
        elif col_idx == 3:
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.font = font_kw
        elif col_idx in [4, 5]:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.font = font_main
        elif col_idx == 6:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if 'P0' in val:
                cell.fill = p0_fill
                cell.font = p0_font
            elif 'P1' in val:
                cell.fill = p1_fill
                cell.font = p1_font
            else:
                cell.fill = p2_fill
                cell.font = p2_font

# Column Widths
col_widths = {
    'A': 10,  # 序号
    'B': 22,  # 关键词分类
    'C': 28,  # 目标搜索关键词
    'D': 38,  # 目标客群画像
    'E': 44,  # 搜索意图与商业价值
    'F': 16   # 推荐优先级
}

for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# Freeze pane below headers
ws.freeze_panes = 'A5'

output_path = r'd:\jinyuliangyuan\今遇莨缘_B端精准引流搜索关键词库.xlsx'
wb.save(output_path)

# Also copy to brain artifacts directory
artifact_dir = r'C:\Users\Administrator\.gemini\antigravity\brain\20658d50-0eed-48c3-8904-ad24bfde3454'
if os.path.exists(artifact_dir):
    artifact_path = os.path.join(artifact_dir, '今遇莨缘_B端精准引流搜索关键词库.xlsx')
    shutil.copy(output_path, artifact_path)
    print("Copied to artifact dir:", artifact_path)

print("Saved successfully to:", output_path)
